#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
transform_swatches.py v3.0 -- Server-Side Bulk Swatch Generator (CORS-free)
============================================================================
Runs entirely server-side, so the browser CORS wall does NOT apply here.
This is the reliable path for catalogs whose image host does not send
Access-Control-Allow-Origin headers (the "CORS blocked" error in the browser
tool). The only thing that can still block a server-side run is the HOST
rate-limiting you (HTTP 429) or requiring authentication (401/403) -- both of
which this script handles explicitly:

  * --delay   : polite pause between requests
  * retries   : automatic exponential backoff, honors the host's Retry-After
  * errors    : 401/403/404/429/timeout are reported distinctly so you know
                if it is throttling (fixable by --delay) vs auth (needs an
                authenticated export).

Speed wins over v2.x:
  - bbox detection on a 256px thumbnail (not a full-image pixel loop)
  - crop computed ONCE per image, .resize() per size (build_base + render_size)
  - Image.draft() fast large-JPEG decode
  - EXIF auto-rotate via ImageOps.exif_transpose
  - ZIP_STORED (images are already compressed)
  - Thread-safe inter-request delay (--delay) serialized across workers

Install:  pip install Pillow requests openpyxl tqdm

Examples:
  python transform_swatches.py -i catalog.xlsx --url-col "Image URL" \
      --sizes 128,256,512 --shape square --format png \
      --layout by-product --concurrency 6 --delay 1 -o swatches.zip

  python transform_swatches.py -i urls.txt --sizes 500,1000 --shape circle

  python transform_swatches.py -i images/ --sizes 300,600 --shape rounded

If you get 429s: raise --delay (e.g. 3) and lower --concurrency (e.g. 3).
If you get 401/403: images need a login -- no delay/proxy will help.
"""
import argparse, concurrent.futures, io, math, os, re, sys, threading, time, zipfile
from pathlib import Path
try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
except ImportError:
    sys.exit("pip install requests")
try:
    from PIL import Image, ImageChops, ImageColor, ImageDraw, ImageOps
except ImportError:
    sys.exit("pip install Pillow")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.exit("pip install openpyxl")
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

VERSION = "3.0.0"
SHAPES = ["square", "circle", "rounded", "pill", "oval", "diamond", "hexagon"]
BG_PRESETS = {"white": (255, 255, 255, 255), "black": (0, 0, 0, 255),
              "transparent": (0, 0, 0, 0), "gray": (128, 128, 128, 255)}
SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}

# Thread-safe inter-request throttle
_delay_lock = threading.Lock()
_last_request_ts = [0.0]


def resolve_bg(value):
    """Accept a named preset OR any hex/CSS color string."""
    if value in BG_PRESETS:
        return BG_PRESETS[value]
    try:
        rgb = ImageColor.getrgb(value)
        return rgb + (255,) if len(rgb) == 3 else rgb
    except Exception:
        return BG_PRESETS["white"]


def make_session(retries=4, backoff=1.5, pool=16):
    s = requests.Session()
    retry = Retry(
        total=retries, connect=retries, read=retries,
        backoff_factor=backoff,
        status_forcelist=(429, 500, 502, 503, 504),
        respect_retry_after_header=True,
        allowed_methods=frozenset(["GET"]),
    )
    a = HTTPAdapter(max_retries=retry, pool_connections=pool, pool_maxsize=pool)
    s.mount("http://", a)
    s.mount("https://", a)
    s.headers.update({
        "User-Agent": "SwatchCLI/" + VERSION,
        "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
    })
    return s


def _throttle(delay):
    """Enforce >= delay seconds between the START of consecutive requests."""
    if delay <= 0:
        return
    with _delay_lock:
        wait = _last_request_ts[0] + delay - time.time()
        if wait > 0:
            time.sleep(wait)
        _last_request_ts[0] = time.time()


def classify_http_error(status):
    if status in (401, 403):
        return (f"HTTP {status} - image requires authentication; "
                f"--delay/proxy will NOT help. Use an authenticated export.")
    if status == 404:
        return "HTTP 404 - image URL not found."
    if status == 429:
        return "HTTP 429 - rate limited even after backoff; raise --delay / lower --concurrency."
    return f"HTTP {status}."


def smart_name(sku, row_name, fallback, idx):
    """SKU -> Name/Title -> filename -> 4-digit numeric"""
    raw = None
    if sku and str(sku).strip():
        raw = str(sku).strip()
    elif row_name and str(row_name).strip():
        raw = str(row_name).strip()
    elif fallback and str(fallback).strip():
        raw = Path(fallback).stem.strip()
    if raw:
        safe = re.sub(r"[^\w\-_. ()]", "_", raw).strip().replace(" ", "_")
        if safe:
            return safe
    return f"swatch_{str(idx + 1).zfill(4)}"


def open_image(buf_or_path, max_target):
    """Open + EXIF-orient. Use Image.draft() to decode big JPEGs at reduced scale."""
    img = Image.open(buf_or_path)
    try:
        img.draft("RGB", (max_target * 2, max_target * 2))
    except Exception:
        pass
    img = ImageOps.exif_transpose(img)
    return img.convert("RGBA")


def download_image(url, session, timeout, delay, max_target):
    _throttle(delay)
    r = session.get(url, timeout=timeout)
    if r.status_code >= 400:
        raise requests.HTTPError(classify_http_error(r.status_code))
    img = open_image(io.BytesIO(r.content), max_target)
    fname = url.split("/")[-1].split("?")[0] or "img.jpg"
    return img, fname


def load_local(path, max_target):
    return open_image(path, max_target), Path(path).name


def detect_bbox(img, threshold=245):
    """Fast bbox of the non-white/non-transparent region via a 256px thumbnail."""
    iw, ih = img.size
    scale = min(1.0, 256.0 / max(iw, ih))
    small = img if scale >= 1.0 else img.resize(
        (max(1, int(iw * scale)), max(1, int(ih * scale))), Image.BILINEAR)
    r, g, b, a = small.split()
    mn = ImageChops.darker(ImageChops.darker(r, g), b)
    color_fg = mn.point(lambda v: 255 if v < threshold else 0)
    alpha_fg = a.point(lambda v: 255 if v >= 128 else 0)
    mask = ImageChops.multiply(color_fg, alpha_fg)
    bb = mask.getbbox()
    if not bb:
        return (0, 0, iw, ih)
    inv = 1.0 / scale
    return (int(bb[0] * inv), int(bb[1] * inv), int(bb[2] * inv), int(bb[3] * inv))


def build_base(img, engine, thresh, pad):
    """Return a square RGBA canvas; per-size output is just .resize((sz,sz))."""
    iw, ih = img.size
    if engine == "stretch":
        return img
    if engine == "contain":
        side = max(iw, ih)
        canvas = Image.new("RGBA", (side, side), (0, 0, 0, 0))
        canvas.paste(img, ((side - iw) // 2, (side - ih) // 2), img)
        return canvas
    if engine == "smart":
        x0, y0, x1, y1 = detect_bbox(img, thresh)
        bw, bh = x1 - x0, y1 - y0
        p = int(max(bw, bh) * pad)
        x0, y0 = max(0, x0 - p), max(0, y0 - p)
        x1, y1 = min(iw, x1 + p), min(ih, y1 + p)
        crop = img.crop((x0, y0, x1, y1))
        cw, ch = crop.size
        side = max(cw, ch)
        canvas = Image.new("RGBA", (side, side), (0, 0, 0, 0))
        canvas.paste(crop, ((side - cw) // 2, (side - ch) // 2), crop)
        return canvas
    sq = min(iw, ih)
    l, t = (iw - sq) // 2, (ih - sq) // 2
    return img.crop((l, t, l + sq, t + sq))


def apply_shape(img, shape):
    w, h = img.size
    mask = Image.new("L", (w, h), 0)
    draw = ImageDraw.Draw(mask)
    if shape == "circle":
        draw.ellipse([0, 0, w - 1, h - 1], fill=255)
    elif shape == "rounded":
        r = max(4, min(w, h) // 8)
        draw.rounded_rectangle([0, 0, w - 1, h - 1], radius=r, fill=255)
    elif shape == "pill":
        r = min(w, h) // 2
        draw.rounded_rectangle([0, 0, w - 1, h - 1], radius=r, fill=255)
    elif shape == "oval":
        draw.ellipse([0, int(h * .1), w, int(h * .9)], fill=255)
    elif shape == "diamond":
        draw.polygon([(w // 2, 0), (w, h // 2), (w // 2, h), (0, h // 2)], fill=255)
    elif shape == "hexagon":
        cx, cy = w // 2, h // 2
        rv = min(cx, cy)
        pts = [(int(cx + rv * math.cos(math.radians(60 * i - 30))),
                int(cy + rv * math.sin(math.radians(60 * i - 30)))) for i in range(6)]
        draw.polygon(pts, fill=255)
    else:
        return img
    res = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    res.paste(img, mask=mask)
    return res


def render_size(base, size, shape, bg, fmt, quality):
    """Resize the pre-built square base to size, apply shape mask, composite bg."""
    sq = base.resize((size, size), Image.LANCZOS)
    shaped = apply_shape(sq, shape)
    out = Image.new("RGBA", (size, size), bg)
    out.paste(shaped, mask=shaped.split()[3])
    buf = io.BytesIO()
    f = fmt.upper()
    if f in ("JPEG", "JPG"):
        out.convert("RGB").save(buf, format="JPEG", quality=quality, optimize=True)
    elif f == "WEBP":
        out.save(buf, format="WEBP", quality=quality)
    else:
        out.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def read_input(path, sheet=None):
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        it = ws.iter_rows(values_only=True)
        hdrs = next(it)
        headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(hdrs)]
        return [dict(zip(headers, r)) for r in it]
    elif ext == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return []
        if rows[0] and re.match(r"^https?://", rows[0][0].strip(), re.I):
            return [{"url": r[0].strip()} for r in rows if r and r[0].strip()]
        headers = rows[0]
        return [dict(zip(headers, r)) for r in rows[1:]]
    elif ext in (".txt", ""):
        with open(path, encoding="utf-8") as f:
            return [{"url": l.strip()} for l in f if l.strip().lower().startswith("http")]
    raise ValueError(f"Unsupported input type: {ext}")


def auto_col(headers, pattern):
    for h in headers:
        if re.search(pattern, str(h), re.I):
            return h
    return None


def _new_result(idx):
    return {"idx": idx, "sku": None, "row_name": None, "url": None,
            "swatches": {}, "status": "ok", "error": None,
            "output_names": {}, "orig_filename": None, "ms": 0}


def _emit_sizes(base, args, res, base_name, bg, ext):
    for sz in args.sizes:
        res["swatches"][sz] = render_size(base, sz, args.shape, bg,
                                          args.output_format, args.quality)
        res["output_names"][sz] = f"{base_name}_{sz}.{ext}"


def process_row(row, idx, session, args):
    res = _new_result(idx)
    t0 = time.time()
    try:
        uc = args.url_col or auto_col(list(row.keys()), r"(image|img|photo|url|src|thumb)")
        url = str(row.get(uc) or "").strip() if uc else ""
        res["url"] = url
        res["sku"] = str(row.get(args.sku_col) or "").strip() if args.sku_col else None
        res["row_name"] = str(row.get(args.name_col) or "").strip() if args.name_col else None
        if not re.match(r"^https?://", url, re.I):
            raise ValueError(f"No valid URL in row {idx + 1}")
        img, fname = download_image(url, session, args.timeout, args.delay, max(args.sizes))
        res["orig_filename"] = fname
        base_name = smart_name(res["sku"], res["row_name"], fname, idx)
        bg = resolve_bg(args.bg_color)
        ext = {"PNG": "png", "JPEG": "jpg", "JPG": "jpg", "WEBP": "webp"}.get(
            args.output_format.upper(), "png")
        base = build_base(img, args.crop_engine, args.bg_threshold, args.pad_pct / 100)
        _emit_sizes(base, args, res, base_name, bg, ext)
    except Exception as e:
        res["status"] = "error"
        res["error"] = str(e)
    res["ms"] = int((time.time() - t0) * 1000)
    return res


def process_folder(folder, args):
    results = []
    bg = resolve_bg(args.bg_color)
    ext = {"PNG": "png", "JPEG": "jpg", "JPG": "jpg", "WEBP": "webp"}.get(
        args.output_format.upper(), "png")
    files = [p for p in Path(folder).rglob("*") if p.suffix.lower() in SUPPORTED_EXTS]
    bar = tqdm(total=len(files), desc="Processing", unit="img") if HAS_TQDM else None
    for i, p in enumerate(files):
        res = _new_result(i)
        res["url"] = str(p)
        res["orig_filename"] = p.name
        t0 = time.time()
        try:
            img = open_image(p, max(args.sizes))
            base_name = smart_name(None, None, p.stem, i)
            base = build_base(img, args.crop_engine, args.bg_threshold, args.pad_pct / 100)
            _emit_sizes(base, args, res, base_name, bg, ext)
        except Exception as e:
            res["status"] = "error"
            res["error"] = str(e)
        res["ms"] = int((time.time() - t0) * 1000)
        results.append(res)
        if bar:
            bar.update(1)
    if bar:
        bar.close()
    return results


def transform_image(img, tw, th, fit="crop", pad_color="#ffffff",
                    fmt="PNG", quality=92, upscale=False):
    """Resize a PIL image to tw x th with the chosen fit mode."""
    sw, sh = img.size
    if not upscale:
        tw = min(tw, sw)
        th = min(th, sh)
    buf = io.BytesIO()
    if fit == "pad":
        try:
            pc = ImageColor.getrgb(pad_color) + (255,)
        except Exception:
            pc = (255, 255, 255, 255)
        r = min(tw / sw, th / sh)
        nw, nh = int(sw * r), int(sh * r)
        rs = img.resize((nw, nh), Image.LANCZOS)
        canvas = Image.new("RGBA", (tw, th), pc)
        canvas.paste(rs, ((tw - nw) // 2, (th - nh) // 2))
        out = canvas
    elif fit == "stretch":
        out = img.resize((tw, th), Image.LANCZOS)
    else:
        r = max(tw / sw, th / sh)
        nw, nh = int(sw * r), int(sh * r)
        rs = img.resize((nw, nh), Image.LANCZOS)
        left, top = (nw - tw) // 2, (nh - th) // 2
        out = rs.crop((left, top, left + tw, top + th))
    f = fmt.upper()
    if f in ("JPEG", "JPG"):
        out.convert("RGB").save(buf, format="JPEG", quality=quality, optimize=True)
    elif f == "WEBP":
        out.save(buf, format="WEBP", quality=quality)
    else:
        out.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def process_row_transform(row, idx, session, args):
    """Process one row in --mode transform."""
    res = _new_result(idx)
    t0 = time.time()
    try:
        uc = args.url_col or auto_col(list(row.keys()), r"(image|img|photo|url|src|thumb)")
        url = str(row.get(uc) or "").strip() if uc else ""
        res["url"] = url
        res["sku"] = str(row.get(args.sku_col) or "").strip() if args.sku_col else None
        res["row_name"] = str(row.get(args.name_col) or "").strip() if args.name_col else None
        if not re.match(r"^https?://", url, re.I):
            raise ValueError(f"No valid URL in row {idx + 1}")
        img, fname = download_image(url, session, args.timeout, args.delay,
                                    max(args.sizes) if args.sizes else 1200)
        res["orig_filename"] = fname
        base_name = smart_name(res["sku"], res["row_name"], fname, idx)
        tw = getattr(args, "transform_width", 1200)
        th = getattr(args, "transform_height", 1200)
        ext = {"PNG": "png", "JPEG": "jpg", "JPG": "jpg", "WEBP": "webp"}.get(
            args.output_format.upper(), "jpg")
        data = transform_image(img, tw, th,
                               fit=getattr(args, "fit", "crop"),
                               pad_color=getattr(args, "pad_color", "#ffffff"),
                               fmt=args.output_format,
                               quality=getattr(args, "quality", 92),
                               upscale=getattr(args, "upscale", False))
        res["swatches"][0] = data
        res["output_names"][0] = f"{base_name}.{ext}"
    except Exception as e:
        res["status"] = "error"
        res["error"] = str(e)
    res["ms"] = int((time.time() - t0) * 1000)
    return res


def write_zip(results, out, layout="flat"):
    ok = [r for r in results if r and r["status"] == "ok"]
    seen = set()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_STORED) as zf:
        for r in ok:
            names = list(r["output_names"].values())
            base = names[0].rsplit("_", 1)[0] if names else f"img_{r['idx'] + 1}"
            for sz, data in r["swatches"].items():
                fn = r["output_names"].get(sz, f"{base}_{sz}.png")
                arc = (fn if layout == "flat" else
                       (f"{sz}/{fn}" if layout == "by-size" else f"{base}/{fn}"))
                while arc in seen:
                    d = arc.rfind(".")
                    arc = (arc[:d] + "_2" + arc[d:]) if d > 0 else arc + "_2"
                seen.add(arc)
                zf.writestr(arc, data)


def write_report(results, out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["#", "Filename", "SKU", "Name", "URL",
               "Status", "Error", "Sizes", "Output Names", "ms"])
    for r in results:
        if not r:
            continue
        ws.append([r["idx"] + 1, r.get("orig_filename", ""), r.get("sku") or "",
                   r.get("row_name") or "", r.get("url") or "", r.get("status", ""),
                   r.get("error") or "",
                   ", ".join(str(s) for s in r.get("swatches", {}).keys()),
                   ", ".join(r.get("output_names", {}).values()), r.get("ms", 0)])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 50)
    wb.save(out)


def parse_args():
    p = argparse.ArgumentParser(
        description="Server-side Bulk Swatch Generator v" + VERSION,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    p.add_argument("--input", "-i", required=True,
                   help="Excel/CSV/TXT URL-list/folder")
    p.add_argument("--out", "-o", default="swatches.zip")
    p.add_argument("--report", default="swatch_report.xlsx")
    p.add_argument("--sizes", "-s", default="128,256,512")
    p.add_argument("--shape", default="square", choices=SHAPES)
    p.add_argument("--crop-engine", default="smart",
                   choices=["smart", "center", "contain", "cover", "stretch"])
    p.add_argument("--bg-color", default="white",
                   help="Named (white/black/transparent/gray) OR hex e.g. #f5f1eb")
    p.add_argument("--bg-threshold", default=245, type=int)
    p.add_argument("--pad-pct", default=5, type=float)
    p.add_argument("--format", default="PNG",
                   choices=["PNG", "JPEG", "WEBP"],
                   type=lambda s: s.upper(), dest="output_format",
                   help="Output format -- case-insensitive: png/PNG/jpeg/JPEG/webp/WEBP")
    p.add_argument("--quality", default=92, type=int,
                   help="JPEG/WebP quality 1-100 (default: 92)")
    p.add_argument("--concurrency", "-c", default=6, type=int)
    p.add_argument("--retries", default=4, type=int)
    p.add_argument("--backoff", default=1.5, type=float,
                   help="Exponential backoff factor (default: 1.5)")
    p.add_argument("--delay", default=0.0, type=float,
                   help="Min seconds between request starts. Raise to 1-10 if you get 429s.")
    p.add_argument("--timeout", default=30, type=int)
    p.add_argument("--sheet", default=None)
    p.add_argument("--url-col", default=None,
                   help="URL column name (auto-detected if omitted)")
    p.add_argument("--sku-col", default=None,
                   help="SKU column for smart naming (auto-detected)")
    p.add_argument("--name-col", default=None,
                   help="Name/Title column for smart naming (auto-detected)")
    p.add_argument("--layout", default="by-product",
                   choices=["flat", "by-product", "by-size"])
    p.add_argument("--mode", default="swatch",
                   choices=["swatch", "transform"])
    p.add_argument("--transform-width", "--tw", default=1200, type=int,
                   dest="transform_width")
    p.add_argument("--transform-height", "--th", default=1200, type=int,
                   dest="transform_height")
    p.add_argument("--fit", default="crop",
                   choices=["crop", "pad", "stretch"])
    p.add_argument("--pad-color", default="#ffffff", dest="pad_color")
    p.add_argument("--upscale", action="store_true")
    p.add_argument("--version", "-v", action="store_true")
    return p.parse_args()


def main():
    args = parse_args()
    if args.version:
        print(f"v{VERSION}")
        return
    args.sizes = [int(s.strip()) for s in args.sizes.split(",") if s.strip().isdigit()]
    if not args.sizes:
        sys.exit("No valid sizes. Use --sizes 128,256,512")

    print(f"\nSwatch Generator v{VERSION} (server-side -- CORS does not apply)")
    print(f"Input       : {args.input}")
    print(f"Sizes       : {args.sizes}")
    print(f"Shape       : {args.shape}")
    print(f"Delay       : {args.delay}s/request  Concurrency: {args.concurrency}\n")

    process_fn = process_row if args.mode == "swatch" else process_row_transform

    if os.path.isdir(args.input):
        results = process_folder(args.input, args)
    else:
        rows = read_input(args.input, sheet=args.sheet)
        if not rows:
            sys.exit("No rows found.")
        hdrs = list(rows[0].keys())
        if not args.url_col:
            args.url_col = auto_col(hdrs, r"(image|img|photo|url|src|thumb)")
            if args.url_col:
                print(f"URL col     : {args.url_col!r}")
        if not args.sku_col:
            args.sku_col = auto_col(hdrs, r"^sku")
            if args.sku_col:
                print(f"SKU col     : {args.sku_col!r}")
        if not args.name_col:
            args.name_col = auto_col(hdrs, r"^(name|title|product)")
            if args.name_col:
                print(f"Name col    : {args.name_col!r}")
        print(f"Rows        : {len(rows)}\n")

        session = make_session(args.retries, args.backoff)
        results = [None] * len(rows)
        bar = tqdm(total=len(rows), desc="Processing", unit="img") if HAS_TQDM else None

        with concurrent.futures.ThreadPoolExecutor(max_workers=args.concurrency) as pool:
            futs = {pool.submit(process_fn, rows[i], i, session, args): i
                    for i in range(len(rows))}
            done = 0
            for fut in concurrent.futures.as_completed(futs):
                i = futs[fut]
                try:
                    results[i] = fut.result()
                except Exception as e:
                    results[i] = _new_result(i)
                    results[i]["status"] = "error"
                    results[i]["error"] = str(e)
                done += 1
                if bar:
                    bar.update(1)
                else:
                    print(f"\r  {done}/{len(rows)}", end="", flush=True)
        if bar:
            bar.close()
        else:
            print()

    ok = sum(1 for r in results if r and r["status"] == "ok")
    err = len(results) - ok
    print(f"\nDone: {ok} OK, {err} errors")

    if err:
        kinds = {}
        for r in results:
            if r and r["status"] == "error":
                key = (r.get("error") or "unknown")[:60]
                kinds[key] = kinds.get(key, 0) + 1
        for k, n in sorted(kinds.items(), key=lambda x: -x[1])[:8]:
            print(f"  {n:>5}x  {k}")

    if ok:
        write_zip(results, args.out, args.layout)
        print(f"\nZIP     : {args.out}")
    write_report(results, args.report)
    print(f"Report  : {args.report}\n")

    if err and not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
